"""

A* grid planning

author: Atsushi Sakai(@Atsushi_twi)
        Nikos Kanargias (nkana@tee.gr)

See Wikipedia article (https://en.wikipedia.org/wiki/A*_search_algorithm)

"""

import sys
import math
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
import multiprocessing

show_animation = False

class AStarPlanner:

    def __init__(self, ox, oy, resolution, rr):
        """
        Initialize grid map for a star planning

        ox: x position list of Obstacles [m]
        oy: y position list of Obstacles [m]
        resolution: grid resolution [m],地图的像素
        rr: robot radius[m]
        """

        self.resolution = resolution
        self.rr = rr
        self.min_x, self.min_y = 0, 0
        self.max_x, self.max_y = 0, 0
        self.obstacle_map = None
        self.x_width, self.y_width = 0, 0
        self.motion = self.get_motion_model()
        self.calc_obstacle_map(ox, oy)

    class Node:
        """定义搜索区域节点类,每个Node都包含坐标x和y, 移动代价cost和父节点索引。
        """
        def __init__(self, x, y, cost, parent_index):
            self.x = int(x)  # index of grid
            self.y = int(y)  # index of grid
            self.cost = cost
            self.parent_index = int(parent_index)

        def __str__(self):
            return str(self.x) + "," + str(self.y) + "," + str(
                self.cost) + "," + str(self.parent_index)

    def planning(self, sx, sy, gx, gy):
        """
        A star path search
        输入起始点和目标点的坐标(sx,sy)和(gx,gy)，
        最终输出的结果是路径包含的点的坐标集合rx和ry。
        input:
            s_x: start x position [m]
            s_y: start y position [m]
            gx: goal x position [m]
            gy: goal y position [m]

        output:
            rx: x position list of the final path
            ry: y position list of the final path
        """
        start_node = self.Node(self.calc_xy_index(sx, self.min_x),
                               self.calc_xy_index(sy, self.min_y), 0.0, -1)
        goal_node = self.Node(self.calc_xy_index(gx, self.min_x),
                              self.calc_xy_index(gy, self.min_y), 0.0, -1)

        open_set, closed_set = dict(), dict()
        open_set[self.calc_grid_index(start_node)] = start_node

        while 1:
            if len(open_set) == 0:
                print("Open set is empty..")
                break

            c_id = min(
                open_set,
                key=lambda o: open_set[o].cost + self.calc_heuristic(goal_node, open_set[o]))
            current = open_set[c_id]

            # show graph
            if show_animation:  # pragma: no cover
                plt.plot(self.calc_grid_position(current.x, self.min_x),
                         self.calc_grid_position(current.y, self.min_y), "xc")
                # for stopping simulation with the esc key.
                plt.gcf().canvas.mpl_connect('key_release_event', lambda event: [exit(0) if event.key == 'escape' else None])
                if len(closed_set.keys()) % 10 == 0:
                    plt.pause(0.001)

            # 通过追踪当前位置current.x和current.y来动态展示路径寻找
            if current.x == goal_node.x and current.y == goal_node.y:
                # print("Find goal")
                goal_node.parent_index = current.parent_index
                goal_node.cost = current.cost
                break

            # Remove the item from the open set
            del open_set[c_id]

            # Add it to the closed set
            closed_set[c_id] = current

            # expand_grid search grid based on motion model
            for i, _ in enumerate(self.motion):
                node = self.Node(current.x + self.motion[i][0],
                                 current.y + self.motion[i][1],
                                 current.cost + self.motion[i][2], c_id)
                n_id = self.calc_grid_index(node)

                # If the node is not safe, do nothing
                if not self.verify_node(node):
                    continue

                if n_id in closed_set:
                    continue

                if n_id not in open_set:
                    open_set[n_id] = node  # discovered a new node
                else:
                    if open_set[n_id].cost > node.cost:
                        # This path is the best until now. record it
                        open_set[n_id] = node

        rx, ry = self.calc_final_path(goal_node, closed_set)

        return rx, ry

    def calc_final_path(self, goal_node, closed_set):
        # generate final course
        rx, ry = [self.calc_grid_position(goal_node.x, self.min_x)], [
            self.calc_grid_position(goal_node.y, self.min_y)]
        parent_index = goal_node.parent_index
        while parent_index != -1:
            n = closed_set[parent_index]
            rx.append(self.calc_grid_position(n.x, self.min_x))
            ry.append(self.calc_grid_position(n.y, self.min_y))
            parent_index = n.parent_index

        return rx, ry

    @staticmethod
    def calc_heuristic(n1, n2):
        """计算启发函数

        Args:
            n1 (_type_): _description_
            n2 (_type_): _description_

        Returns:
            _type_: _description_
        """
        w = 1.0  # weight of heuristic
        d = w * math.hypot(n1.x - n2.x, n1.y - n2.y)
        return d

    def calc_grid_position(self, index, min_position):
        """
        calc grid position

        :param index:
        :param min_position:
        :return:
        """
        pos = index * self.resolution + min_position
        return pos

    def calc_xy_index(self, position, min_pos):
        return round((position - min_pos) / self.resolution)

    def calc_grid_index(self, node):
        return (node.y - self.min_y) * self.x_width + (node.x - self.min_x)

    def verify_node(self, node):
        px = self.calc_grid_position(node.x, self.min_x)
        py = self.calc_grid_position(node.y, self.min_y)

        if px < self.min_x:
            return False
        elif py < self.min_y:
            return False
        elif px >= self.max_x:
            return False
        elif py >= self.max_y:
            return False

        # collision check
        if self.obstacle_map[node.x][node.y]:
            return False

        return True

    def calc_obstacle_map(self, ox, oy):

        self.min_x = round(min(ox))
        self.min_y = round(min(oy))
        self.max_x = round(max(ox))
        self.max_y = round(max(oy))
        # print("min_x:", self.min_x)
        # print("min_y:", self.min_y)
        # print("max_x:", self.max_x)
        # print("max_y:", self.max_y)

        self.x_width = int(round((self.max_x - self.min_x) / self.resolution))
        self.y_width = int(round((self.max_y - self.min_y) / self.resolution))
        # print("x_width:", self.x_width)
        # print("y_width:", self.y_width)

        # obstacle map generation
        self.obstacle_map = [[False for _ in range(self.y_width)]
                             for _ in range(self.x_width)]
        for ix in range(self.x_width):
            x = self.calc_grid_position(ix, self.min_x)
            for iy in range(self.y_width):
                y = self.calc_grid_position(iy, self.min_y)
                for iox, ioy in zip(ox, oy):
                    d = math.hypot(iox - x, ioy - y)
                    if d <= self.rr:
                        self.obstacle_map[ix][iy] = True
                        break

    @staticmethod
    def get_motion_model():
        # dx, dy, cost
        motion = [[1, 0, 1],
                  [0, 1, 1],
                  [-1, 0, 1],
                  [0, -1, 1],
                  [-1, -1, math.sqrt(2)],
                  [-1, 1, math.sqrt(2)],
                  [1, -1, math.sqrt(2)],
                  [1, 1, math.sqrt(2)],
                  [2, 1, math.sqrt(5)],
                  [2, -1, math.sqrt(5)],
                  [-2, 1, math.sqrt(5)],
                  [-2, -1, math.sqrt(5)],
                  [1, 2, math.sqrt(5)],
                  [1, -2, math.sqrt(5)],
                  [-1, 2, math.sqrt(5)],
                  [-1, -2, math.sqrt(5)],]

        return motion

def square(center, edge, interval):
    """
    生成方形障碍，center为正方形中心，edge为正方形边长，interval为障碍点间隔。
    """
    ox = []
    oy = []

    for i in range(int(edge/interval)):
        ox.append(center[0]-edge/2)
        oy.append(center[1]-edge/2+i*interval)

        ox.append(center[0]+edge/2)
        oy.append(center[1]-edge/2+(i+1)*interval)

        oy.append(center[1]+edge/2)
        ox.append(center[0]-edge/2+i*interval)

        oy.append(center[1]-edge/2)
        ox.append(center[0]-edge/2+(i+1)*interval)

    return ox, oy

def circle(center, radius, interval):
    """
    生成圆形障碍，center为圆心，radius/2为半径，interval为障碍点间隔。
    """
    ox = []
    oy = []

    circumference = 2 * np.pi * radius
    number = int(circumference/interval)

    for i in range(number):
        ox.append(center[0] + np.cos(i / number * 2 * np.pi) * radius / 2)
        oy.append(center[1] + np.sin(i / number * 2 * np.pi) * radius / 2)

    return ox, oy

def main(sx, sy, gx, gy, grid_size, robot_radius, obstacle, edge):
    ox, oy = [], []
    for i in range(edge[0], edge[1]+1):
        ox.append(i)
        oy.append(edge[2])
        ox.append(i)
        oy.append(edge[3])
    for i in range(edge[2], edge[3] + 1):
        ox.append(edge[0])
        oy.append(i)
        ox.append(edge[1])
        oy.append(i)

    for i in range(obstacle.shape[0]):
        if(obstacle[i, -1]==0.0):
            ox_sta, oy_sta = circle([obstacle[i, 0], obstacle[i, 1]], obstacle[i, 2], 0.1)
            ox = ox + ox_sta
            oy = oy + oy_sta
        else:
            ox_sta, oy_sta = square([obstacle[i, 0], obstacle[i, 1]], obstacle[i, 2], 0.1)
            ox = ox + ox_sta
            oy = oy + oy_sta

    if show_animation:  # pragma: no cover
        plt.plot(ox, oy, ".k")
        plt.plot(sx, sy, "og")
        plt.plot(gx, gy, "xb")
        plt.grid(True)
        plt.axis("equal")

    a_star = AStarPlanner(ox, oy, grid_size, robot_radius)
    rx, ry = a_star.planning(sx, sy, gx, gy)

    if show_animation:  # pragma: no cover
        plt.plot(rx, ry, "-r")

    coords = np.hstack([np.array(rx).reshape(-1, 1), np.array(ry).reshape(-1, 1)])
    coords = collinear(coords)
    if show_animation:  # pragma: no cover
        plt.plot([i[0] for i in coords], [i[1] for i in coords], color='red', marker='*', markersize=10)
    coords = collinear(coords)
    if show_animation:  # pragma: no cover
        plt.plot([i[0] for i in coords], [i[1] for i in coords], color='green', marker='*', markersize=10)
        plt.pause(0.001)
        plt.show()
    return coords

def triangle_area(A, B, C):
    """
    计算由三个点A(x1, y1), B(x2, y2), C(x3, y3)构成的三角形的面积，面积为0则说明共线。
    """
    x1 = A[0]; x2 = B[0]; x3 = C[0]; y1 = A[1]; y2 = B[1]; y3 = C[1]
    # 使用行列式方法计算面积
    # 面积 = 0.5 * |x1(y2-y3) + x2(y3-y1) + x3(y1-y2)|
    area = 0.5 * abs(x1 * (y2 - y3) + x2 * (y3 - y1) + x3 * (y1 - y2))
    return area

def collinear(point):
    """
    清除point中不影响形状的点。
    """
    index = []
    i = 0
    while (i != (point.shape[0] - 2)):
        for j in range(i + 2, point.shape[0]):
            area = triangle_area(point[i, :], point[i + 1, :], point[j, :])
            if (area == 0.0):
                index.append(j - 1)
            else:
                break
        i = j - 1
    return np.delete(point, index, axis=0)

def plan_path(args):
    index1, index2, obj1, obj2, grid_size, agent_radius, obstacles, x_min, x_max, y_min, y_max = args
    sx = obj1[0]
    sy = obj1[1]
    gx = obj2[0]
    gy = obj2[1]

    coords = main(
        sx=sx, sy=sy, gx=gx, gy=gy,
        grid_size=grid_size,
        robot_radius=agent_radius,
        obstacle=obstacles,
        edge=[math.floor(x_min), math.ceil(x_max), math.floor(y_min), math.ceil(y_max)]
    )

    path_len = 0.0
    for i in range(coords.shape[0] - 1):
        path_len += np.sqrt((coords[i + 1, 0] - coords[i, 0]) ** 2 + (coords[i + 1, 1] - coords[i, 1]) ** 2)

    data = coords
    reversed_data = data[::-1]
    data = np.vstack([reversed_data, data[1:, :]])
    data = pd.DataFrame(data)
    return index1, index2, data, path_len

def create_or_clear_excel(excel_path):
    # 确保路径中所有目录都存在
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    if not os.path.exists(excel_path):
        wb = Workbook()
    else:
        wb = load_workbook(excel_path)
        # 删除所有现有的 sheet
        for sheet in wb.sheetnames:
            std = wb[sheet]
            wb.remove(std)
        # 添加一个干净的空 sheet
        wb.create_sheet("Sheet1")

    wb.save(excel_path)


if __name__ == '__main__':
    '''Environment location'''
    script_dir = os.path.dirname(os.path.abspath(__file__))
    all_sheets = pd.read_excel(script_dir + r"\excel_file\env.xlsx", sheet_name=None, header=None)
    obstacles = all_sheets["obstacles"].to_numpy()
    opponents = all_sheets["opponents"].to_numpy()
    base = all_sheets["base"].to_numpy()[0]
    grid_size = 1.0  # 网格大小
    agent_radius = 1.5  # 机器人大小

    excel_path = script_dir + r"\excel_file\flight_point.xlsx"
    create_or_clear_excel(excel_path)

    '''场景边界'''
    x_min, x_max, y_min, y_max = base[0] - 5, base[0] + 5, base[1] - 5, base[1] + 5
    for index in range(obstacles.shape[0]):
        obs = obstacles[index, :]
        x_min = min(x_min, obs[0] - obs[2] / 2 - 5)
        x_max = max(x_max, obs[0] + obs[2] / 2 + 5)
        y_min = min(y_min, obs[1] - obs[2] / 2 - 5)
        y_max = max(y_max, obs[1] + obs[2] / 2 + 5)

    args_list = [
        (0, i + 1, base, opponents[i, :], grid_size, agent_radius, obstacles, x_min, x_max, y_min, y_max)
        for i in range(opponents.shape[0])
    ]

    with multiprocessing.Pool(processes=min(opponents.shape[0], multiprocessing.cpu_count())) as pool:
        result = pool.map(plan_path, args_list)

    writer = pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace', engine='openpyxl')
    for res in result:
        res[2].to_excel(writer, sheet_name=str("path from " + str(res[0]) + " to " + str(res[1])), header=None, index=False)

    path_len = np.zeros([opponents.shape[0], opponents.shape[0]])
    path_len_sub = np.array([res[3] for res in result])
    path_len[0, :] = path_len_sub

    obstacles = np.vstack([obstacles, base])
    for i in range(opponents.shape[0]-1):
        print(i)
        args_list = [
            (i + 1, j + 1, opponents[i, :], opponents[j, :], grid_size, agent_radius, obstacles, x_min, x_max, y_min, y_max)
            for j in range(i + 1, opponents.shape[0])
        ]

        with multiprocessing.Pool(processes=min(opponents.shape[0], multiprocessing.cpu_count())) as pool:
            result = pool.map(plan_path, args_list)

        for res in result:
            res[2].to_excel(writer, sheet_name=str("path from " + str(res[0]) + " to " + str(res[1])), header=None, index=False)
        path_len_sub = np.array([res[3] for res in result])
        path_len[i+1, (i+1):] = path_len_sub
    path_len = pd.DataFrame(path_len)
    path_len.to_excel(writer, sheet_name=str("path_len"), header=None, index=False)
    writer.save()
    writer.close()