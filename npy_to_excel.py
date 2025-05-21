import numpy as np
import pandas as pd
import os
import sys
import parameter
from openpyxl import Workbook, load_workbook

strategies = ["Quantity-based", "AST", "TASE", "prediction-driven", "prediction-driven-NN", "prediction-driven-NN_deception", "TNNLS"]
npy_names = ["act_seq", "allies_s1", "opponent_exc", "opponent_rec", "opponent_exc_vel", "opponent_rec_vel", "opponents"]
# strategies = ["Quantity-based"]
# npy_names = ["opponent_exc"]

for strategy in strategies:
    print(strategy)
    folder_path = sys.path[0] + "\\data\\" + strategy + "\\"
    excel_path = sys.path[0] + "\\data\\" + strategy + ".xlsx"
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    if not os.path.exists(excel_path):
        wb = Workbook()
    else:
        wb = load_workbook(excel_path)
        # 删除所有现有的 sheet
        for sheet in wb.sheetnames:
            std = wb[sheet]
            wb.remove(std)
        wb.create_sheet("Sheet1")
    wb.save(excel_path)

    for npy in npy_names:
        data_sequence = []
        data_merge = np.zeros([20000, 20000])
        rows = 0; columns = 0
        for phase in range(len(parameter.P)):
            npy_path = folder_path + npy + "_Adv_phase=" + str(phase) + "_KE=" + str(parameter.KE) \
                       + "_KR=" + str(parameter.KR) + "_KO=" + str(parameter.KO) + ".npy"
            if os.path.exists(npy_path):
                data = np.load(npy_path, allow_pickle=True)
                if(npy == "allies_s1"):
                    data = data.T

                if (npy == "opponent_exc" or npy == "opponent_exc_vel"):
                    if (np.all(data == 0, axis=0)[0] == 0):
                        data = data[1:, :]

                zero_rows = np.all(data == 0, axis=1)
                zero_row_indices = np.where(zero_rows)[0]
                columns = max(columns, data.shape[1])
                # print(columns)
                if zero_row_indices.size == 0:
                    data_merge[rows:(data.shape[0]+rows), :min(columns, data.shape[1])] = data
                    rows += data.shape[0]
                else:
                    data_merge[rows:(zero_row_indices[0] + rows), :min(columns, data.shape[1])] = data[:zero_row_indices[0], :]
                    rows += zero_row_indices[0]
            else:
                None
                # print(f"文件不存在: {npy_path}")
        zero_rows = np.all(data_merge == 0, axis=1)
        zero_row_indices = np.where(zero_rows)[0]
        data_merge = data_merge[:zero_row_indices[0], :columns]
        df = pd.DataFrame(data_merge)

        with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=npy, index=False, header=False)
sys.exit()






# 读取文件名
files = os.listdir(folder_path)

# 根据类型分类
type_to_files = {}

for file in files:
    if file.endswith(".npy"):
        # 提取类型名，比如 act_seq_Adv
        type_name = file.split("_phase=")[0]
        type_to_files.setdefault(type_name, []).append(file)

# 开始处理
with pd.ExcelWriter(save_path) as writer:
    for type_name, type_files in type_to_files.items():
        # 按phase排序
        type_files.sort(key=lambda x: int(x.split("_phase=")[1].split(" ")[0]))

        arrays = []
        for filename in type_files:
            data = np.load(os.path.join(folder_path, filename))
            arrays.append(data)

        # 拼接
        concatenated = np.concatenate(arrays, axis=0)

        # 转成DataFrame保存
        df = pd.DataFrame(concatenated)
        df.to_excel(writer, sheet_name=type_name[:30], index=False)  # sheet_name最长31字符
